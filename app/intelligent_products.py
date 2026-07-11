from __future__ import annotations

import json
import sqlite3
from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from openpyxl import load_workbook

from app import database


router = APIRouter(prefix="/products/intelligent", tags=["intelligent-products"])

IMPORT_HEADERS = [
    "codigo_interno",
    "descricao",
    "categoria",
    "marca",
    "unidade",
    "fornecedor",
    "fabricante",
    "referencia_fabricante",
    "ncm",
    "ipi",
    "icms",
    "codigo_barras",
    "codigo_erp",
    "sistema_erp",
    "empresa_codigo",
    "status",
    "estoque_minimo",
    "estoque_maximo",
    "lead_time_dias",
    "localizacao",
    "custo_medio",
    "ultimo_custo",
    "ultima_compra",
    "estoque_atual",
    "estoque_reservado",
]


def _main():
    import app.main as main

    return main


def _user(request: Request, permission: str):
    main = _main()
    user = main.require(request, permission)
    return user


def _context(request: Request, user, **extra):
    return _main().base_context(request, user, **extra)


def _templates():
    return _main().templates


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _code(value: Any) -> str:
    return _clean(value).upper()


def _float(value: Any, field: str, errors: list[str]) -> float:
    if value in (None, ""):
        return 0.0
    try:
        number = float(value)
        if number < 0:
            errors.append(f"{field}: não pode ser negativo")
        return number
    except (TypeError, ValueError):
        errors.append(f"{field}: valor numérico inválido")
        return 0.0


def _int(value: Any, field: str, errors: list[str]) -> int:
    if value in (None, ""):
        return 0
    try:
        number = int(float(value))
        if number < 0:
            errors.append(f"{field}: não pode ser negativo")
        return number
    except (TypeError, ValueError):
        errors.append(f"{field}: número inteiro inválido")
        return 0


def _date(value: Any, errors: list[str]) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = _clean(value)
    try:
        return datetime.fromisoformat(text[:10]).strftime("%Y-%m-%d")
    except ValueError:
        errors.append("ultima_compra: use AAAA-MM-DD")
        return None


def _find_id(
    conn: sqlite3.Connection,
    table: str,
    value: Any,
    *,
    column: str = "name",
) -> int | None:
    text = _clean(value)
    if not text:
        return None
    row = conn.execute(
        f"SELECT id FROM {table} WHERE {column}=? COLLATE NOCASE",
        (text,),
    ).fetchone()
    return row["id"] if row else None


def _supplier_id(conn, value: Any) -> int | None:
    text = _clean(value)
    if not text:
        return None
    row = conn.execute(
        """
        SELECT id FROM suppliers
        WHERE code=? COLLATE NOCASE
           OR legal_name=? COLLATE NOCASE
           OR trade_name=? COLLATE NOCASE
        """,
        (text, text, text),
    ).fetchone()
    return row["id"] if row else None


def _manufacturer_id(conn, value: Any) -> int | None:
    text = _clean(value)
    if not text:
        return None
    # A Sprint 3 pode existir em bancos atualizados; em bases sem a tabela,
    # a ausência é tratada como cadastro inválido.
    try:
        row = conn.execute(
            """
            SELECT id FROM manufacturers
            WHERE code=? COLLATE NOCASE OR name=? COLLATE NOCASE
            """,
            (text, text),
        ).fetchone()
        return row["id"] if row else None
    except sqlite3.OperationalError:
        return None


def _duplicate_matches(
    conn: sqlite3.Connection,
    *,
    internal_code: str = "",
    erp_code: str = "",
    barcode: str = "",
    manufacturer_reference: str = "",
    exclude_id: int | None = None,
) -> list[dict[str, Any]]:
    clauses = []
    params: list[Any] = []

    fields = {
        "internal_code": internal_code,
        "erp_code": erp_code,
        "barcode": barcode,
        "manufacturer_reference": manufacturer_reference,
    }
    for field, value in fields.items():
        if value:
            clauses.append(f"p.{field}=? COLLATE NOCASE")
            params.append(value)

    if not clauses:
        return []

    sql = f"""
        SELECT p.id, p.internal_code, p.description, p.erp_code,
               p.barcode, p.manufacturer_reference
        FROM products p
        WHERE ({' OR '.join(clauses)})
    """
    if exclude_id:
        sql += " AND p.id<>?"
        params.append(exclude_id)

    return [dict(row) for row in conn.execute(sql, params).fetchall()]


def _validate_row(conn, row_number: int, values: dict[str, Any]) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []

    internal_code = _code(values.get("codigo_interno"))
    description = _clean(values.get("descricao"))
    unit_code = _code(values.get("unidade"))
    erp_code = _clean(values.get("codigo_erp"))
    barcode = _clean(values.get("codigo_barras"))
    manufacturer_reference = _clean(values.get("referencia_fabricante"))

    if not internal_code:
        errors.append("codigo_interno: obrigatório")
    if not description:
        errors.append("descricao: obrigatória")
    if not unit_code:
        errors.append("unidade: obrigatória")

    unit_id = _find_id(conn, "units", unit_code, column="code")
    if unit_code and not unit_id:
        errors.append(f"unidade: '{unit_code}' não cadastrada")

    category_text = _clean(values.get("categoria"))
    category_id = _find_id(conn, "categories", category_text)
    if category_text and not category_id:
        errors.append(f"categoria: '{category_text}' não cadastrada")

    brand_text = _clean(values.get("marca"))
    brand_id = _find_id(conn, "brands", brand_text)
    if brand_text and not brand_id:
        errors.append(f"marca: '{brand_text}' não cadastrada")

    supplier_text = _clean(values.get("fornecedor"))
    supplier_id = _supplier_id(conn, supplier_text)
    if supplier_text and not supplier_id:
        errors.append(f"fornecedor: '{supplier_text}' não cadastrado")

    manufacturer_text = _clean(values.get("fabricante"))
    manufacturer_id = _manufacturer_id(conn, manufacturer_text)
    if manufacturer_text and not manufacturer_id:
        errors.append(f"fabricante: '{manufacturer_text}' não cadastrado")

    company_code = _clean(values.get("empresa_codigo"))
    company_id = None
    if company_code:
        company = conn.execute(
            "SELECT id FROM companies WHERE code=? AND active=1",
            (company_code,),
        ).fetchone()
        if company:
            company_id = company["id"]
        else:
            errors.append(f"empresa_codigo: '{company_code}' não cadastrada")

    duplicates = _duplicate_matches(
        conn,
        internal_code=internal_code,
        erp_code=erp_code,
        barcode=barcode,
        manufacturer_reference=manufacturer_reference,
    )

    existing = conn.execute(
        "SELECT id FROM products WHERE internal_code=? COLLATE NOCASE",
        (internal_code,),
    ).fetchone() if internal_code else None

    operation = "UPDATE" if existing else "CREATE"

    for match in duplicates:
        if existing and match["id"] == existing["id"]:
            continue
        conflicting = []
        if internal_code and match["internal_code"] == internal_code:
            conflicting.append("código interno")
        if erp_code and match["erp_code"] == erp_code:
            conflicting.append("código ERP")
        if barcode and match["barcode"] == barcode:
            conflicting.append("código de barras")
        if (
            manufacturer_reference
            and match["manufacturer_reference"] == manufacturer_reference
        ):
            conflicting.append("referência do fabricante")
        if conflicting:
            errors.append(
                f"duplicidade em {', '.join(conflicting)} com "
                f"{match['internal_code']} — {match['description']}"
            )

    ipi = _float(values.get("ipi"), "ipi", errors)
    icms = _float(values.get("icms"), "icms", errors)
    minimum_stock = _float(
        values.get("estoque_minimo"), "estoque_minimo", errors
    )
    maximum_stock = _float(
        values.get("estoque_maximo"), "estoque_maximo", errors
    )
    if maximum_stock and maximum_stock < minimum_stock:
        errors.append("estoque_maximo: deve ser maior ou igual ao mínimo")

    payload = {
        "internal_code": internal_code,
        "description": description,
        "category_id": category_id,
        "brand_id": brand_id,
        "unit_id": unit_id,
        "default_supplier_id": supplier_id,
        "manufacturer_id": manufacturer_id,
        "manufacturer_reference": manufacturer_reference or None,
        "ncm": _clean(values.get("ncm")) or None,
        "ipi_rate": ipi,
        "icms_rate": icms,
        "barcode": barcode or None,
        "erp_code": erp_code or None,
        "active": 0 if _code(values.get("status")) == "INATIVO" else 1,
        "source_system": _code(values.get("sistema_erp")) or "MANUAL",
        "company_id": company_id,
        "external_description": description,
        "minimum_stock": minimum_stock,
        "maximum_stock": maximum_stock,
        "lead_time_days": _int(
            values.get("lead_time_dias"), "lead_time_dias", errors
        ),
        "stock_location": _clean(values.get("localizacao")) or None,
        "average_cost": _float(
            values.get("custo_medio"), "custo_medio", errors
        ),
        "last_cost": _float(
            values.get("ultimo_custo"), "ultimo_custo", errors
        ),
        "last_purchase_date": _date(values.get("ultima_compra"), errors),
        "current_stock": _float(
            values.get("estoque_atual"), "estoque_atual", errors
        ),
        "reserved_stock": _float(
            values.get("estoque_reservado"), "estoque_reservado", errors
        ),
    }

    if erp_code and not company_id:
        warnings.append(
            "codigo_erp informado sem empresa; o código externo será global"
        )

    return {
        "row_number": row_number,
        "operation": operation,
        "internal_code": internal_code,
        "description": description,
        "payload": payload,
        "errors": errors,
        "warnings": warnings,
    }


@router.get("/search")
def intelligent_search(
    request: Request,
    term: str = Query(..., min_length=2, max_length=100),
    limit: int = Query(12, ge=1, le=30),
):
    user = _user(request, "catalog.smart_search")
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")

    text = term.strip()
    like = f"%{text}%"
    with database.connect() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT
                p.id,
                p.internal_code,
                p.description,
                p.erp_code,
                p.barcode,
                p.manufacturer_reference,
                pec.source_system,
                pec.external_code
            FROM products p
            LEFT JOIN product_external_codes pec ON pec.product_id=p.id
            WHERE p.internal_code LIKE ? COLLATE NOCASE
               OR p.description LIKE ? COLLATE NOCASE
               OR p.erp_code LIKE ? COLLATE NOCASE
               OR p.barcode LIKE ? COLLATE NOCASE
               OR p.manufacturer_reference LIKE ? COLLATE NOCASE
               OR pec.external_code LIKE ? COLLATE NOCASE
            ORDER BY
                CASE WHEN p.internal_code=? COLLATE NOCASE THEN 0 ELSE 1 END,
                p.description
            LIMIT ?
            """,
            (like, like, like, like, like, like, text, limit),
        ).fetchall()

    return JSONResponse(
        {
            "term": text,
            "results": [dict(row) for row in rows],
        }
    )


@router.get("/duplicates")
def duplicate_check(
    request: Request,
    internal_code: str = "",
    erp_code: str = "",
    barcode: str = "",
    manufacturer_reference: str = "",
    exclude_id: int | None = None,
):
    user = _user(request, "catalog.smart_search")
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")

    with database.connect() as conn:
        matches = _duplicate_matches(
            conn,
            internal_code=_code(internal_code),
            erp_code=_clean(erp_code),
            barcode=_clean(barcode),
            manufacturer_reference=_clean(manufacturer_reference),
            exclude_id=exclude_id,
        )

    return {"duplicate": bool(matches), "matches": matches}


@router.post("/import/validate", response_class=HTMLResponse)
async def validate_import(
    request: Request,
    file: UploadFile = File(...),
):
    user = _user(request, "catalog.import")
    if not user:
        return _main().login_redirect()

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse(
            "/products?import_error=formato",
            status_code=303,
        )

    workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    sheet = workbook.active
    headers = [
        _clean(cell.value) for cell in sheet[1]
    ]

    missing = [header for header in IMPORT_HEADERS if header not in headers]
    if missing:
        return _templates().TemplateResponse(
            "import_validation.html",
            _context(
                request,
                user,
                batch=None,
                rows=[],
                header_errors=[
                    "Colunas ausentes: " + ", ".join(missing)
                ],
            ),
            status_code=422,
        )

    token = database.create_import_token()
    validated_rows = []

    with database.connect() as conn:
        for number, raw in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            values = dict(zip(headers, raw))
            if not any(values.get(header) not in (None, "") for header in headers):
                continue
            validated_rows.append(_validate_row(conn, number, values))

        valid_count = sum(not row["errors"] for row in validated_rows)
        error_count = len(validated_rows) - valid_count
        cursor = conn.execute(
            """
            INSERT INTO product_import_batches(
                token, filename, user_id, status,
                total_rows, valid_rows, error_rows
            ) VALUES(?, ?, ?, 'VALIDATED', ?, ?, ?)
            """,
            (
                token,
                file.filename,
                user["id"],
                len(validated_rows),
                valid_count,
                error_count,
            ),
        )
        batch_id = cursor.lastrowid

        for row in validated_rows:
            conn.execute(
                """
                INSERT INTO product_import_rows(
                    batch_id, row_number, operation, internal_code,
                    payload_json, errors_json, warnings_json
                ) VALUES(?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    row["row_number"],
                    row["operation"],
                    row["internal_code"],
                    database.json_dump(row["payload"]),
                    database.json_dump(row["errors"]),
                    database.json_dump(row["warnings"]),
                ),
            )
        conn.commit()

    database.audit(
        user["id"],
        "product.import.validated",
        f"batch={batch_id};validas={valid_count};erros={error_count}",
    )

    return _templates().TemplateResponse(
        "import_validation.html",
        _context(
            request,
            user,
            batch={
                "id": batch_id,
                "token": token,
                "filename": file.filename,
                "total_rows": len(validated_rows),
                "valid_rows": valid_count,
                "error_rows": error_count,
            },
            rows=validated_rows,
            header_errors=[],
        ),
    )


def _apply_product(conn, payload: dict[str, Any], operation: str) -> int:
    existing = conn.execute(
        "SELECT id FROM products WHERE internal_code=? COLLATE NOCASE",
        (payload["internal_code"],),
    ).fetchone()

    product_values = (
        payload["description"],
        payload["category_id"],
        payload["brand_id"],
        payload["unit_id"],
        payload["default_supplier_id"],
        payload["ncm"],
        payload["ipi_rate"],
        payload["icms_rate"],
        payload["barcode"],
        payload["erp_code"],
        payload["active"],
        payload["manufacturer_id"],
        payload["manufacturer_reference"],
    )

    if existing:
        product_id = existing["id"]
        conn.execute(
            """
            UPDATE products SET
                description=?,
                category_id=?,
                brand_id=?,
                unit_id=?,
                default_supplier_id=?,
                ncm=?,
                ipi_rate=?,
                icms_rate=?,
                barcode=?,
                erp_code=?,
                active=?,
                manufacturer_id=?,
                manufacturer_reference=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            product_values + (product_id,),
        )
    else:
        cursor = conn.execute(
            """
            INSERT INTO products(
                internal_code, description, category_id, brand_id,
                unit_id, default_supplier_id, ncm, ipi_rate, icms_rate,
                barcode, erp_code, active, manufacturer_id,
                manufacturer_reference
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (payload["internal_code"],) + product_values,
        )
        product_id = cursor.lastrowid

    if payload["company_id"]:
        conn.execute(
            """
            INSERT INTO product_company_settings(
                product_id, company_id, minimum_stock, maximum_stock,
                lead_time_days, stock_location, average_cost, last_cost,
                last_purchase_date, current_stock, reserved_stock, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(product_id, company_id) DO UPDATE SET
                minimum_stock=excluded.minimum_stock,
                maximum_stock=excluded.maximum_stock,
                lead_time_days=excluded.lead_time_days,
                stock_location=excluded.stock_location,
                average_cost=excluded.average_cost,
                last_cost=excluded.last_cost,
                last_purchase_date=excluded.last_purchase_date,
                current_stock=excluded.current_stock,
                reserved_stock=excluded.reserved_stock,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                product_id,
                payload["company_id"],
                payload["minimum_stock"],
                payload["maximum_stock"],
                payload["lead_time_days"],
                payload["stock_location"],
                payload["average_cost"],
                payload["last_cost"],
                payload["last_purchase_date"],
                payload["current_stock"],
                payload["reserved_stock"],
            ),
        )

    if payload["erp_code"]:
        conn.execute(
            """
            INSERT INTO product_external_codes(
                product_id, source_system, company_id, external_code,
                external_description, sync_status, updated_at
            ) VALUES(?, ?, ?, ?, ?, 'IMPORTED', CURRENT_TIMESTAMP)
            ON CONFLICT(source_system, company_id, external_code)
            DO UPDATE SET
                product_id=excluded.product_id,
                external_description=excluded.external_description,
                sync_status='IMPORTED',
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                product_id,
                payload["source_system"],
                payload["company_id"],
                payload["erp_code"],
                payload["external_description"],
            ),
        )

    return product_id


@router.post("/import/{batch_id}/confirm")
def confirm_import(
    request: Request,
    batch_id: int,
    token: str = Form(...),
):
    user = _user(request, "catalog.import.confirm")
    if not user:
        return _main().login_redirect()

    with database.connect() as conn:
        batch = conn.execute(
            """
            SELECT * FROM product_import_batches
            WHERE id=? AND token=?
            """,
            (batch_id, token),
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="Lote inexistente")
        if batch["status"] != "VALIDATED":
            raise HTTPException(
                status_code=409,
                detail="Este lote já foi processado",
            )
        if batch["error_rows"] > 0:
            raise HTTPException(
                status_code=422,
                detail="Corrija os erros antes de confirmar",
            )

        rows = conn.execute(
            """
            SELECT * FROM product_import_rows
            WHERE batch_id=? ORDER BY row_number
            """,
            (batch_id,),
        ).fetchall()

        applied = 0
        try:
            conn.execute("BEGIN")
            for row in rows:
                errors = json.loads(row["errors_json"])
                if errors:
                    continue
                payload = json.loads(row["payload_json"])
                product_id = _apply_product(
                    conn,
                    payload,
                    row["operation"],
                )
                _main().save_history(
                    conn,
                    product_id,
                    user["id"],
                    f"smart_import.{row['operation'].lower()}",
                )
                conn.execute(
                    """
                    UPDATE product_import_rows
                    SET applied_product_id=?
                    WHERE id=?
                    """,
                    (product_id, row["id"]),
                )
                applied += 1

            conn.execute(
                """
                UPDATE product_import_batches
                SET status='CONFIRMED', confirmed_at=CURRENT_TIMESTAMP
                WHERE id=?
                """,
                (batch_id,),
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise

    database.audit(
        user["id"],
        "product.import.confirmed",
        f"batch={batch_id};aplicadas={applied}",
    )
    return RedirectResponse(
        f"/products?import_success={applied}",
        status_code=303,
    )


@router.get("/item/{product_id}/external-codes", response_class=HTMLResponse)
def external_codes_page(request: Request, product_id: int):
    user = _user(request, "catalog.read")
    if not user:
        return _main().login_redirect()

    with database.connect() as conn:
        product = conn.execute(
            "SELECT * FROM products WHERE id=?",
            (product_id,),
        ).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="Produto inexistente")
        codes = conn.execute(
            """
            SELECT pec.*, c.code company_code,
                   COALESCE(c.trade_name, c.legal_name) company_name
            FROM product_external_codes pec
            LEFT JOIN companies c ON c.id=pec.company_id
            WHERE pec.product_id=?
            ORDER BY pec.source_system, c.code, pec.external_code
            """,
            (product_id,),
        ).fetchall()
        companies = conn.execute(
            """
            SELECT id, code, legal_name, trade_name
            FROM companies WHERE active=1 ORDER BY code
            """
        ).fetchall()

    return _templates().TemplateResponse(
        "product_external_codes.html",
        _context(
            request,
            user,
            product=product,
            codes=codes,
            companies=companies,
        ),
    )


@router.post("/item/{product_id}/external-codes")
def add_external_code(
    request: Request,
    product_id: int,
    source_system: str = Form(...),
    external_code: str = Form(...),
    company_id: int | None = Form(None),
    external_description: str = Form(""),
):
    user = _user(request, "catalog.external_codes")
    if not user:
        return _main().login_redirect()

    try:
        with database.connect() as conn:
            product = conn.execute(
                "SELECT id FROM products WHERE id=?",
                (product_id,),
            ).fetchone()
            if not product:
                raise HTTPException(
                    status_code=404,
                    detail="Produto inexistente",
                )
            conn.execute(
                """
                INSERT INTO product_external_codes(
                    product_id, source_system, company_id, external_code,
                    external_description, sync_status
                ) VALUES(?, ?, ?, ?, ?, 'MANUAL')
                """,
                (
                    product_id,
                    _code(source_system),
                    company_id,
                    _clean(external_code),
                    _clean(external_description) or None,
                ),
            )
            conn.commit()
    except sqlite3.IntegrityError:
        return RedirectResponse(
            f"/products/intelligent/item/{product_id}/external-codes"
            "?error=duplicate",
            status_code=303,
        )

    database.audit(
        user["id"],
        "product.external_code.created",
        f"produto={product_id};sistema={_code(source_system)}",
    )
    return RedirectResponse(
        f"/products/intelligent/item/{product_id}/external-codes",
        status_code=303,
    )
