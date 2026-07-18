from __future__ import annotations

import csv
import importlib
import json
import os
import sqlite3
import urllib.request
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.integration.models import ConnectionTestResult, ReadResult


class BaseConnector(ABC):
    connector_type: str
    version = "1.0.0"

    @abstractmethod
    def test_connection(
        self,
        config: dict[str, Any],
        secrets: dict[str, str],
    ) -> ConnectionTestResult:
        raise NotImplementedError

    @abstractmethod
    def read(
        self,
        config: dict[str, Any],
        secrets: dict[str, str],
        cursor: str | None = None,
        limit: int = 10000,
    ) -> ReadResult:
        raise NotImplementedError


class CSVConnector(BaseConnector):
    connector_type = "CSV"

    def _path(self, config: dict[str, Any]) -> Path:
        path = Path(str(config.get("path", ""))).expanduser().resolve()
        if not path.exists() or not path.is_file():
            raise FileNotFoundError(f"Arquivo não encontrado: {path}")
        return path

    def test_connection(self, config, secrets):
        try:
            path = self._path(config)
            delimiter = str(config.get("delimiter", ";"))
            encoding = str(config.get("encoding", "utf-8-sig"))
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle, delimiter=delimiter)
                headers = reader.fieldnames or []
            return ConnectionTestResult(
                True,
                f"Arquivo acessível com {len(headers)} coluna(s).",
                {"headers": headers, "size_bytes": path.stat().st_size},
            )
        except Exception as exc:
            return ConnectionTestResult(False, str(exc))

    def read(self, config, secrets, cursor=None, limit=10000):
        path = self._path(config)
        delimiter = str(config.get("delimiter", ";"))
        encoding = str(config.get("encoding", "utf-8-sig"))
        start = int(cursor or 0)
        rows = []
        with path.open("r", encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            for index, row in enumerate(reader):
                if index < start:
                    continue
                rows.append(dict(row))
                if len(rows) >= limit:
                    return ReadResult(
                        rows=rows,
                        next_cursor=str(index + 1),
                    )
        return ReadResult(rows=rows, next_cursor=None)


class ExcelConnector(BaseConnector):
    connector_type = "EXCEL"

    def _load(self, config):
        path = Path(str(config.get("path", ""))).expanduser().resolve()
        if not path.exists() or not path.is_file():
            raise FileNotFoundError(f"Arquivo não encontrado: {path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = config.get("sheet")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        return path, workbook, sheet

    def test_connection(self, config, secrets):
        try:
            path, workbook, sheet = self._load(config)
            headers = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]
            workbook.close()
            return ConnectionTestResult(
                True,
                f"Planilha '{sheet.title}' acessível.",
                {
                    "headers": headers,
                    "sheet": sheet.title,
                    "size_bytes": path.stat().st_size,
                },
            )
        except Exception as exc:
            return ConnectionTestResult(False, str(exc))

    def read(self, config, secrets, cursor=None, limit=10000):
        _, workbook, sheet = self._load(config)
        iterator = sheet.iter_rows(values_only=True)
        headers = [
            str(value).strip() if value is not None else ""
            for value in next(iterator)
        ]
        start = int(cursor or 0)
        rows = []
        for index, values in enumerate(iterator):
            if index < start:
                continue
            rows.append(dict(zip(headers, values)))
            if len(rows) >= limit:
                workbook.close()
                return ReadResult(
                    rows=rows,
                    next_cursor=str(index + 1),
                )
        workbook.close()
        return ReadResult(rows=rows, next_cursor=None)


class RESTConnector(BaseConnector):
    connector_type = "REST"

    def _request(self, config, secrets):
        url = str(config.get("url", "")).strip()
        if not url.startswith(("http://", "https://")):
            raise ValueError("URL inválida")
        headers = {"Accept": "application/json"}
        token = secrets.get("TOKEN")
        if token:
            headers["Authorization"] = f"Bearer {token}"
        request = urllib.request.Request(url, headers=headers, method="GET")
        return urllib.request.urlopen(request, timeout=10)

    def test_connection(self, config, secrets):
        try:
            with self._request(config, secrets) as response:
                return ConnectionTestResult(
                    200 <= response.status < 300,
                    f"HTTP {response.status}",
                    {"content_type": response.headers.get("Content-Type")},
                )
        except Exception as exc:
            return ConnectionTestResult(False, str(exc))

    def read(self, config, secrets, cursor=None, limit=10000):
        with self._request(config, secrets) as response:
            payload = json.loads(response.read().decode("utf-8"))
        path = str(config.get("records_path", "")).strip()
        data = payload
        if path:
            for part in path.split("."):
                data = data[part]
        if not isinstance(data, list):
            raise ValueError("A resposta configurada não é uma lista")
        rows = [dict(item) for item in data[:limit]]
        return ReadResult(rows=rows)


class OptionalDatabaseConnector(BaseConnector):
    module_name: str
    install_hint: str

    def _module(self):
        try:
            return importlib.import_module(self.module_name)
        except ImportError as exc:
            raise RuntimeError(
                f"Driver opcional ausente. Instale: {self.install_hint}"
            ) from exc

    def test_connection(self, config, secrets):
        try:
            module = self._module()
            connection = self._connect(module, config, secrets)
            cursor = connection.cursor()
            cursor.execute(self._validation_query())
            cursor.fetchone()
            cursor.close()
            connection.close()
            return ConnectionTestResult(True, "Conexão somente leitura validada.")
        except Exception as exc:
            return ConnectionTestResult(False, str(exc))

    def read(self, config, secrets, cursor=None, limit=10000):
        module = self._module()
        query = str(config.get("query", "")).strip()
        if not query.lower().startswith("select"):
            raise ValueError("Somente consultas SELECT são permitidas")
        connection = self._connect(module, config, secrets)
        db_cursor = connection.cursor()
        db_cursor.execute(query)
        columns = [item[0] for item in db_cursor.description]
        rows = [
            dict(zip(columns, values))
            for values in db_cursor.fetchmany(limit)
        ]
        db_cursor.close()
        connection.close()
        return ReadResult(rows=rows)

    def _validation_query(self):
        return "SELECT 1"

    @abstractmethod
    def _connect(self, module, config, secrets):
        raise NotImplementedError


class PostgreSQLConnector(OptionalDatabaseConnector):
    connector_type = "POSTGRESQL"
    module_name = "psycopg"
    install_hint = "pip install psycopg[binary]"

    def _connect(self, module, config, secrets):
        return module.connect(
            host=config["host"],
            port=int(config.get("port", 5432)),
            dbname=config["database"],
            user=secrets["USER"],
            password=secrets["PASSWORD"],
            connect_timeout=8,
            options="-c default_transaction_read_only=on",
        )


class MySQLConnector(OptionalDatabaseConnector):
    connector_type = "MYSQL"
    module_name = "mysql.connector"
    install_hint = "pip install mysql-connector-python"

    def _connect(self, module, config, secrets):
        connection = module.connect(
            host=config["host"],
            port=int(config.get("port", 3306)),
            database=config["database"],
            user=secrets["USER"],
            password=secrets["PASSWORD"],
            connection_timeout=8,
        )
        connection.start_transaction(readonly=True)
        return connection


class SQLServerConnector(OptionalDatabaseConnector):
    connector_type = "SQLSERVER"
    module_name = "pyodbc"
    install_hint = "pip install pyodbc"

    def _connect(self, module, config, secrets):
        driver = config.get("driver", "ODBC Driver 18 for SQL Server")
        connection_string = (
            f"DRIVER={{{driver}}};SERVER={config['host']},"
            f"{int(config.get('port', 1433))};DATABASE={config['database']};"
            f"UID={secrets['USER']};PWD={secrets['PASSWORD']};"
            "Encrypt=yes;TrustServerCertificate=yes;"
            "ApplicationIntent=ReadOnly;Connection Timeout=8"
        )
        return module.connect(connection_string)


class OracleConnector(OptionalDatabaseConnector):
    connector_type = "ORACLE"
    module_name = "oracledb"
    install_hint = "pip install oracledb"

    def _connect(self, module, config, secrets):
        dsn = module.makedsn(
            config["host"],
            int(config.get("port", 1521)),
            service_name=config["database"],
        )
        return module.connect(
            user=secrets["USER"],
            password=secrets["PASSWORD"],
            dsn=dsn,
        )

    def _validation_query(self):
        return "SELECT 1 FROM dual"


class FirebirdConnector(OptionalDatabaseConnector):
    connector_type = "FIREBIRD"
    module_name = "firebird.driver"
    install_hint = "pip install firebird-driver"

    def _connect(self, module, config, secrets):
        return module.connect(
            database=config["database"],
            host=config["host"],
            port=int(config.get("port", 3050)),
            user=secrets["USER"],
            password=secrets["PASSWORD"],
        )

    def _validation_query(self):
        return "SELECT 1 FROM RDB$DATABASE"


connector_registry = {
    connector.connector_type: connector
    for connector in [
        CSVConnector(),
        ExcelConnector(),
        RESTConnector(),
        PostgreSQLConnector(),
        MySQLConnector(),
        SQLServerConnector(),
        OracleConnector(),
        FirebirdConnector(),
    ]
}


def load_secrets(prefix: str | None) -> dict[str, str]:
    if not prefix:
        return {}
    normalized = prefix.strip().upper()
    keys = ["USER", "PASSWORD", "TOKEN", "API_KEY"]
    return {
        key: os.environ[f"{normalized}_{key}"]
        for key in keys
        if os.environ.get(f"{normalized}_{key}")
    }
