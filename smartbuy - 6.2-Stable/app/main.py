from contextlib import asynccontextmanager
from datetime import date, datetime
from io import BytesIO
import json
import math

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app import database
from app.security import hash_password, read_session, sign_session, verify_password

BASE=database.BASE; COOKIE='smartbuy_session'

@asynccontextmanager
async def lifespan(_:FastAPI): database.init_db(); yield
app=FastAPI(title='SmartBuy',version='2.0.0',lifespan=lifespan)
app.mount('/static',StaticFiles(directory=BASE/'app'/'static'),name='static')
templates=Jinja2Templates(directory=BASE/'app'/'templates')

def current_user(request):
    uid=read_session(request.cookies.get(COOKIE)); user=database.get_user(uid) if uid else None
    return user if user and user['active'] else None

def allowed(user,p): return p in database.permissions(user['id'])
def login_redirect(): return RedirectResponse('/',status_code=303)
def require(request,p=None):
    user=current_user(request)
    if not user: return None
    if p and not allowed(user,p): raise HTTPException(403,'Sem permissão')
    return user

def base_context(request,user,**extra):
    return {'request':request,'user':user,'permissions':database.permissions(user['id']),**extra}

@app.get('/health')
def health(): return {'status':'ok'}
@app.get('/',response_class=HTMLResponse)
def login_page(request:Request):
    if current_user(request): return RedirectResponse('/dashboard',303)
    return templates.TemplateResponse('login.html',{'request':request,'error':None})
@app.post('/login')
def login(request:Request,email:str=Form(...),password:str=Form(...)):
    user=database.get_user_by_email(email)
    if not user or not user['active'] or not verify_password(password,user['password_hash']):
        return templates.TemplateResponse('login.html',{'request':request,'error':'E-mail ou senha inválidos.'},status_code=401)
    database.audit(user['id'],'auth.login'); response=RedirectResponse('/dashboard',303)
    response.set_cookie(COOKIE,sign_session(user['id']),httponly=True,samesite='lax',max_age=28800); return response
@app.post('/logout')
def logout(request:Request):
    user=current_user(request)
    if user: database.audit(user['id'],'auth.logout')
    response=RedirectResponse('/',303); response.delete_cookie(COOKIE); return response

@app.get('/dashboard',response_class=HTMLResponse)
def dashboard(request:Request):
    user=require(request)
    if not user:return login_redirect()
    with database.connect() as c:
        counts={name:c.execute(sql).fetchone()['total'] for name,sql in {
          'companies':'SELECT COUNT(*) total FROM companies WHERE active=1',
          'users':'SELECT COUNT(*) total FROM users WHERE active=1',
          'products':'SELECT COUNT(*) total FROM products WHERE active=1',
          'suppliers':'SELECT COUNT(*) total FROM suppliers WHERE active=1'}.items()}
    return templates.TemplateResponse('dashboard.html',base_context(request,user,**counts))

# Rotas preservadas da Sprint 1
@app.get('/companies',response_class=HTMLResponse)
def companies(request:Request):
    user=require(request,'companies.read')
    if not user:return login_redirect()
    with database.connect() as c: rows=c.execute('SELECT * FROM companies ORDER BY code').fetchall()
    return templates.TemplateResponse('companies.html',base_context(request,user,companies=rows,error=None))
@app.post('/companies')
def create_company(request:Request,code:str=Form(...),legal_name:str=Form(...),trade_name:str=Form(''),tax_id:str=Form(...),city:str=Form(''),state:str=Form('')):
    user=require(request,'companies.write')
    if not user:return login_redirect()
    try:
        with database.connect() as c:
            c.execute('INSERT INTO companies(code,legal_name,trade_name,tax_id,city,state) VALUES(?,?,?,?,?,?)',(code.strip(),legal_name.strip(),trade_name.strip() or None,tax_id.strip(),city.strip() or None,state.strip().upper() or None)); c.commit()
        database.audit(user['id'],'company.created',code.strip()); return RedirectResponse('/companies',303)
    except Exception:
        with database.connect() as c: rows=c.execute('SELECT * FROM companies ORDER BY code').fetchall()
        return templates.TemplateResponse('companies.html',base_context(request,user,companies=rows,error='Código ou identificação fiscal já cadastrados.'),status_code=409)
@app.get('/users',response_class=HTMLResponse)
def users(request:Request):
    user=require(request,'users.read')
    if not user:return login_redirect()
    with database.connect() as c:
        rows=c.execute('SELECT users.*,roles.name role_name FROM users JOIN roles ON roles.id=users.role_id ORDER BY users.full_name').fetchall(); roles=c.execute('SELECT * FROM roles ORDER BY name').fetchall()
    return templates.TemplateResponse('users.html',base_context(request,user,users=rows,roles=roles,error=None))
@app.post('/users')
def create_user(request:Request,full_name:str=Form(...),email:str=Form(...),password:str=Form(...),role_id:int=Form(...)):
    user=require(request,'users.write')
    if not user:return login_redirect()
    error='A senha deve ter pelo menos 10 caracteres.' if len(password)<10 else None
    if not error:
        try:
            with database.connect() as c:
                c.execute('INSERT INTO users(full_name,email,password_hash,role_id) VALUES(?,?,?,?)',(full_name.strip(),email.strip().lower(),hash_password(password),role_id)); c.commit()
            database.audit(user['id'],'user.created',email.strip().lower()); return RedirectResponse('/users',303)
        except Exception:error='E-mail já cadastrado ou perfil inválido.'
    with database.connect() as c:
        rows=c.execute('SELECT users.*,roles.name role_name FROM users JOIN roles ON roles.id=users.role_id ORDER BY users.full_name').fetchall(); roles=c.execute('SELECT * FROM roles ORDER BY name').fetchall()
    return templates.TemplateResponse('users.html',base_context(request,user,users=rows,roles=roles,error=error),status_code=409)
@app.get('/audit',response_class=HTMLResponse)
def audit(request:Request):
    user=require(request,'audit.read')
    if not user:return login_redirect()
    with database.connect() as c: rows=c.execute('SELECT audit_log.*,users.full_name actor FROM audit_log LEFT JOIN users ON users.id=audit_log.user_id ORDER BY audit_log.id DESC LIMIT 300').fetchall()
    return templates.TemplateResponse('audit.html',base_context(request,user,events=rows))

# Cadastros auxiliares
@app.get('/catalog/settings',response_class=HTMLResponse)
def catalog_settings(request:Request):
    user=require(request,'catalog.read')
    if not user:return login_redirect()
    with database.connect() as c:
        context={k:c.execute(f'SELECT * FROM {table} ORDER BY {order}').fetchall() for k,table,order in [
          ('categories','categories','name'),('brands','brands','name'),('units','units','code'),('suppliers','suppliers','legal_name')]}
    return templates.TemplateResponse('catalog_settings.html',base_context(request,user,error=None,**context))
@app.post('/catalog/settings/{kind}')
def create_setting(request:Request,kind:str,name:str=Form(''),code:str=Form(''),description:str=Form(''),legal_name:str=Form(''),trade_name:str=Form(''),tax_id:str=Form(''),contact_name:str=Form(''),email:str=Form(''),phone:str=Form('')):
    user=require(request,'catalog.write')
    if not user:return login_redirect()
    try:
        with database.connect() as c:
            if kind=='category': c.execute('INSERT INTO categories(name) VALUES(?)',(name.strip(),))
            elif kind=='brand': c.execute('INSERT INTO brands(name) VALUES(?)',(name.strip(),))
            elif kind=='unit': c.execute('INSERT INTO units(code,description) VALUES(?,?)',(code.strip().upper(),description.strip()))
            elif kind=='supplier': c.execute('INSERT INTO suppliers(code,legal_name,trade_name,tax_id,contact_name,email,phone) VALUES(?,?,?,?,?,?,?)',(code.strip(),legal_name.strip(),trade_name.strip() or None,tax_id.strip() or None,contact_name.strip() or None,email.strip() or None,phone.strip() or None))
            else: raise ValueError('Tipo inválido')
            c.commit()
        database.audit(user['id'],f'{kind}.created',name or legal_name or code); return RedirectResponse('/catalog/settings',303)
    except Exception:
        return RedirectResponse('/catalog/settings?error=duplicate',303)

PRODUCT_SELECT='''SELECT p.*,c.name category_name,b.name brand_name,u.code unit_code,s.trade_name supplier_trade_name,s.legal_name supplier_legal_name FROM products p LEFT JOIN categories c ON c.id=p.category_id LEFT JOIN brands b ON b.id=p.brand_id JOIN units u ON u.id=p.unit_id LEFT JOIN suppliers s ON s.id=p.default_supplier_id'''

def product_snapshot(conn,pid):
    row=conn.execute(PRODUCT_SELECT+' WHERE p.id=?',(pid,)).fetchone(); return dict(row) if row else {}
def save_history(conn,pid,uid,action):
    conn.execute('INSERT INTO product_history(product_id,user_id,action,snapshot_json) VALUES(?,?,?,?)',(pid,uid,action,json.dumps(product_snapshot(conn,pid),ensure_ascii=False,default=str)))

def catalog_options(conn):
    return {
      'categories':conn.execute('SELECT * FROM categories WHERE active=1 ORDER BY name').fetchall(),
      'brands':conn.execute('SELECT * FROM brands WHERE active=1 ORDER BY name').fetchall(),
      'units':conn.execute('SELECT * FROM units WHERE active=1 ORDER BY code').fetchall(),
      'suppliers':conn.execute('SELECT * FROM suppliers WHERE active=1 ORDER BY legal_name').fetchall(),
      'companies':conn.execute('SELECT * FROM companies WHERE active=1 ORDER BY code').fetchall(),
    }

@app.get('/products',response_class=HTMLResponse)
def products(request:Request,q:str='',category_id:int|None=None,brand_id:int|None=None,supplier_id:int|None=None,status:str='active',sort:str='description',direction:str='asc',page:int=Query(1,ge=1),per_page:int=Query(20,ge=10,le=100)):
    user=require(request,'catalog.read')
    if not user:return login_redirect()
    allowed_sorts={'internal_code':'p.internal_code','description':'p.description','category':'c.name','brand':'b.name','updated':'p.updated_at'}
    order=allowed_sorts.get(sort,'p.description'); direction='desc' if direction=='desc' else 'asc'
    where=[]; params=[]
    if q:
        where.append('(p.internal_code LIKE ? OR p.description LIKE ? OR p.erp_code LIKE ? OR p.barcode LIKE ?)'); term=f'%{q.strip()}%'; params += [term]*4
    if category_id: where.append('p.category_id=?'); params.append(category_id)
    if brand_id: where.append('p.brand_id=?'); params.append(brand_id)
    if supplier_id: where.append('p.default_supplier_id=?'); params.append(supplier_id)
    if status=='active': where.append('p.active=1')
    elif status=='inactive': where.append('p.active=0')
    clause=' WHERE '+' AND '.join(where) if where else ''
    with database.connect() as c:
        total=c.execute('SELECT COUNT(*) total FROM products p'+clause,params).fetchone()['total']; pages=max(1,math.ceil(total/per_page)); page=min(page,pages)
        rows=c.execute(PRODUCT_SELECT+clause+f' ORDER BY {order} {direction},p.id LIMIT ? OFFSET ?',params+[per_page,(page-1)*per_page]).fetchall(); opts=catalog_options(c)
    return templates.TemplateResponse('products.html',base_context(request,user,products=rows,total=total,page=page,pages=pages,per_page=per_page,q=q,category_id=category_id,brand_id=brand_id,supplier_id=supplier_id,status=status,sort=sort,direction=direction,**opts))

@app.get('/products/new',response_class=HTMLResponse)
def new_product(request:Request):
    user=require(request,'catalog.write')
    if not user:return login_redirect()
    with database.connect() as c: opts=catalog_options(c)
    return templates.TemplateResponse('product_form.html',base_context(request,user,product=None,error=None,**opts))

@app.post('/products/new')
def create_product(request:Request,internal_code:str=Form(...),description:str=Form(...),category_id:int|None=Form(None),brand_id:int|None=Form(None),unit_id:int=Form(...),default_supplier_id:int|None=Form(None),ncm:str=Form(''),ipi_rate:float=Form(0),icms_rate:float=Form(0),barcode:str=Form(''),erp_code:str=Form(''),active:int=Form(1),company_id:int|None=Form(None),minimum_stock:float=Form(0),maximum_stock:float=Form(0),lead_time_days:int=Form(0),stock_location:str=Form(''),average_cost:float=Form(0),last_cost:float=Form(0),last_purchase_date:str=Form(''),current_stock:float=Form(0)):
    user=require(request,'catalog.write')
    if not user:return login_redirect()
    try:
        with database.connect() as c:
            cur=c.execute('''INSERT INTO products(internal_code,description,category_id,brand_id,unit_id,default_supplier_id,ncm,ipi_rate,icms_rate,barcode,erp_code,active) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',(internal_code.strip(),description.strip(),category_id,brand_id,unit_id,default_supplier_id,ncm.strip() or None,ipi_rate,icms_rate,barcode.strip() or None,erp_code.strip() or None,1 if active else 0)); pid=cur.lastrowid
            if company_id: c.execute('''INSERT INTO product_company_settings(product_id,company_id,minimum_stock,maximum_stock,lead_time_days,stock_location,average_cost,last_cost,last_purchase_date,current_stock) VALUES(?,?,?,?,?,?,?,?,?,?)''',(pid,company_id,minimum_stock,maximum_stock,lead_time_days,stock_location.strip() or None,average_cost,last_cost,last_purchase_date or None,current_stock))
            save_history(c,pid,user['id'],'created'); c.commit()
        database.audit(user['id'],'product.created',internal_code.strip()); return RedirectResponse(f'/products/item/{pid}',303)
    except Exception as exc:
        with database.connect() as c: opts=catalog_options(c)
        return templates.TemplateResponse('product_form.html',base_context(request,user,product=None,error='Código interno, ERP ou código de barras já cadastrado.',**opts),status_code=409)

@app.get('/products/item/{product_id}',response_class=HTMLResponse)
def product_detail(request:Request,product_id:int):
    user=require(request,'catalog.read')
    if not user:return login_redirect()
    with database.connect() as c:
        product=c.execute(PRODUCT_SELECT+' WHERE p.id=?',(product_id,)).fetchone()
        if not product: raise HTTPException(404)
        settings=c.execute('''SELECT pcs.*,co.code company_code,co.trade_name company_trade_name,co.legal_name company_legal_name,(pcs.current_stock-pcs.reserved_stock) available_stock FROM product_company_settings pcs JOIN companies co ON co.id=pcs.company_id WHERE pcs.product_id=? ORDER BY co.code''',(product_id,)).fetchall()
        opts=catalog_options(c)
    return templates.TemplateResponse('product_detail.html',base_context(request,user,product=product,settings=settings,**opts))

@app.post('/products/item/{product_id}/stock')
def save_stock(request:Request,product_id:int,company_id:int=Form(...),minimum_stock:float=Form(0),maximum_stock:float=Form(0),lead_time_days:int=Form(0),stock_location:str=Form(''),average_cost:float=Form(0),last_cost:float=Form(0),last_purchase_date:str=Form(''),current_stock:float=Form(0),reserved_stock:float=Form(0)):
    user=require(request,'inventory.write')
    if not user:return login_redirect()
    with database.connect() as c:
        c.execute('''INSERT INTO product_company_settings(product_id,company_id,minimum_stock,maximum_stock,lead_time_days,stock_location,average_cost,last_cost,last_purchase_date,current_stock,reserved_stock,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP) ON CONFLICT(product_id,company_id) DO UPDATE SET minimum_stock=excluded.minimum_stock,maximum_stock=excluded.maximum_stock,lead_time_days=excluded.lead_time_days,stock_location=excluded.stock_location,average_cost=excluded.average_cost,last_cost=excluded.last_cost,last_purchase_date=excluded.last_purchase_date,current_stock=excluded.current_stock,reserved_stock=excluded.reserved_stock,updated_at=CURRENT_TIMESTAMP''',(product_id,company_id,minimum_stock,maximum_stock,lead_time_days,stock_location.strip() or None,average_cost,last_cost,last_purchase_date or None,current_stock,reserved_stock)); save_history(c,product_id,user['id'],'stock.updated'); c.commit()
    database.audit(user['id'],'product.stock.updated',f'produto={product_id};empresa={company_id}'); return RedirectResponse(f'/products/item/{product_id}',303)

@app.get('/products/item/{product_id}/history',response_class=HTMLResponse)
def product_history(request:Request,product_id:int):
    user=require(request,'history.read')
    if not user:return login_redirect()
    with database.connect() as c:
        product=c.execute('SELECT * FROM products WHERE id=?',(product_id,)).fetchone(); rows=c.execute('''SELECT h.*,u.full_name actor FROM product_history h LEFT JOIN users u ON u.id=h.user_id WHERE h.product_id=? ORDER BY h.id DESC''',(product_id,)).fetchall()
    return templates.TemplateResponse('product_history.html',base_context(request,user,product=product,history=rows))

EXPORT_HEADERS=['codigo_interno','descricao','categoria','marca','unidade','fornecedor','ncm','ipi','icms','codigo_barras','codigo_erp','status','empresa_codigo','estoque_minimo','estoque_maximo','lead_time_dias','localizacao','custo_medio','ultimo_custo','ultima_compra','estoque_atual','estoque_reservado']

def style_sheet(ws):
    fill=PatternFill('solid',fgColor='0B5B44')
    for cell in ws[1]: cell.fill=fill; cell.font=Font(color='FFFFFF',bold=True); cell.alignment=Alignment(horizontal='center')
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    widths=[18,38,20,18,12,24,14,10,10,18,18,12,16,16,16,16,18,16,16,16,16,18]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i) if i<=26 else 'A'].width=w

@app.get('/products/export.xlsx')
def export_products(request:Request,q:str='',category_id:int|None=None,brand_id:int|None=None,supplier_id:int|None=None,status:str='all'):
    user=require(request,'catalog.export')
    if not user:return login_redirect()
    where=[]; params=[]
    if q: where.append('(p.internal_code LIKE ? OR p.description LIKE ? OR p.erp_code LIKE ?)'); term=f'%{q}%'; params += [term]*3
    if category_id: where.append('p.category_id=?');params.append(category_id)
    if brand_id:where.append('p.brand_id=?');params.append(brand_id)
    if supplier_id:where.append('p.default_supplier_id=?');params.append(supplier_id)
    if status=='active':where.append('p.active=1')
    elif status=='inactive':where.append('p.active=0')
    clause=' WHERE '+' AND '.join(where) if where else ''
    sql='''SELECT p.*,cat.name category_name,b.name brand_name,u.code unit_code,s.legal_name supplier_name,co.code company_code,pcs.minimum_stock,pcs.maximum_stock,pcs.lead_time_days,pcs.stock_location,pcs.average_cost,pcs.last_cost,pcs.last_purchase_date,pcs.current_stock,pcs.reserved_stock FROM products p LEFT JOIN categories cat ON cat.id=p.category_id LEFT JOIN brands b ON b.id=p.brand_id JOIN units u ON u.id=p.unit_id LEFT JOIN suppliers s ON s.id=p.default_supplier_id LEFT JOIN product_company_settings pcs ON pcs.product_id=p.id LEFT JOIN companies co ON co.id=pcs.company_id'''+clause+' ORDER BY p.description,co.code'
    with database.connect() as c: rows=c.execute(sql,params).fetchall()
    wb=Workbook();ws=wb.active;ws.title='Produtos';ws.append(EXPORT_HEADERS)
    for r in rows: ws.append([r['internal_code'],r['description'],r['category_name'],r['brand_name'],r['unit_code'],r['supplier_name'],r['ncm'],r['ipi_rate'],r['icms_rate'],r['barcode'],r['erp_code'],'ATIVO' if r['active'] else 'INATIVO',r['company_code'],r['minimum_stock'],r['maximum_stock'],r['lead_time_days'],r['stock_location'],r['average_cost'],r['last_cost'],r['last_purchase_date'],r['current_stock'],r['reserved_stock']])
    style_sheet(ws); output=BytesIO();wb.save(output);output.seek(0);database.audit(user['id'],'product.exported',f'{len(rows)} linhas')
    return StreamingResponse(output,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=smartbuy_produtos.xlsx'})

@app.get('/products/template.xlsx')
def product_template(request:Request):
    user=require(request,'catalog.import')
    if not user:return login_redirect()
    wb=Workbook();ws=wb.active;ws.title='Produtos';ws.append(EXPORT_HEADERS);style_sheet(ws);output=BytesIO();wb.save(output);output.seek(0)
    return StreamingResponse(output,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=modelo_importacao_smartbuy.xlsx'})

@app.post('/products/import')
async def import_products(request:Request,file:UploadFile=File(...)):
    user=require(request,'catalog.import')
    if not user:return login_redirect()
    if not file.filename or not file.filename.lower().endswith('.xlsx'): return RedirectResponse('/products?import_error=formato',303)
    data=await file.read(); wb=load_workbook(BytesIO(data),data_only=True);ws=wb.active
    headers=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    if headers[:len(EXPORT_HEADERS)]!=EXPORT_HEADERS: return RedirectResponse('/products?import_error=cabecalho',303)
    success=0;errors=[]
    with database.connect() as c:
        def id_by_name(table,name,code_col='name'):
            if not name:return None
            row=c.execute(f'SELECT id FROM {table} WHERE {code_col}=? COLLATE NOCASE',(str(name).strip(),)).fetchone();return row['id'] if row else None
        for number,row in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
            if not row[0] and not row[1]:continue
            values=dict(zip(EXPORT_HEADERS,row))
            try:
                unit_id=id_by_name('units',values['unidade'],'code')
                if not unit_id: raise ValueError('unidade inexistente')
                category_id=id_by_name('categories',values['categoria']);brand_id=id_by_name('brands',values['marca'])
                supplier_id=None
                if values['fornecedor']:
                    found=c.execute('SELECT id FROM suppliers WHERE legal_name=? COLLATE NOCASE OR trade_name=? COLLATE NOCASE',(str(values['fornecedor']).strip(),str(values['fornecedor']).strip())).fetchone();supplier_id=found['id'] if found else None
                existing=c.execute('SELECT id FROM products WHERE internal_code=?',(str(values['codigo_interno']).strip(),)).fetchone()
                payload=(str(values['descricao']).strip(),category_id,brand_id,unit_id,supplier_id,str(values['ncm']).strip() if values['ncm'] else None,float(values['ipi'] or 0),float(values['icms'] or 0),str(values['codigo_barras']).strip() if values['codigo_barras'] else None,str(values['codigo_erp']).strip() if values['codigo_erp'] else None,0 if str(values['status']).upper()=='INATIVO' else 1)
                if existing:
                    pid=existing['id'];c.execute('''UPDATE products SET description=?,category_id=?,brand_id=?,unit_id=?,default_supplier_id=?,ncm=?,ipi_rate=?,icms_rate=?,barcode=?,erp_code=?,active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?''',payload+(pid,));action='import.updated'
                else:
                    cur=c.execute('''INSERT INTO products(internal_code,description,category_id,brand_id,unit_id,default_supplier_id,ncm,ipi_rate,icms_rate,barcode,erp_code,active) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',(str(values['codigo_interno']).strip(),)+payload);pid=cur.lastrowid;action='import.created'
                if values['empresa_codigo']:
                    company=c.execute('SELECT id FROM companies WHERE code=?',(str(values['empresa_codigo']).strip(),)).fetchone()
                    if not company:raise ValueError('empresa inexistente')
                    c.execute('''INSERT INTO product_company_settings(product_id,company_id,minimum_stock,maximum_stock,lead_time_days,stock_location,average_cost,last_cost,last_purchase_date,current_stock,reserved_stock,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP) ON CONFLICT(product_id,company_id) DO UPDATE SET minimum_stock=excluded.minimum_stock,maximum_stock=excluded.maximum_stock,lead_time_days=excluded.lead_time_days,stock_location=excluded.stock_location,average_cost=excluded.average_cost,last_cost=excluded.last_cost,last_purchase_date=excluded.last_purchase_date,current_stock=excluded.current_stock,reserved_stock=excluded.reserved_stock,updated_at=CURRENT_TIMESTAMP''',(pid,company['id'],float(values['estoque_minimo'] or 0),float(values['estoque_maximo'] or 0),int(values['lead_time_dias'] or 0),str(values['localizacao']).strip() if values['localizacao'] else None,float(values['custo_medio'] or 0),float(values['ultimo_custo'] or 0),str(values['ultima_compra'])[:10] if values['ultima_compra'] else None,float(values['estoque_atual'] or 0),float(values['estoque_reservado'] or 0)))
                save_history(c,pid,user['id'],action);success+=1
            except Exception as exc: errors.append(f'Linha {number}: {exc}')
        c.commit()
    database.audit(user['id'],'product.imported',f'sucesso={success};erros={len(errors)}'); return templates.TemplateResponse('import_result.html',base_context(request,user,success=success,errors=errors))

# SMARTBUY_SPRINT_5_ROUTER
from app.purchasing_intelligence import router as purchasing_intelligence_router

app.include_router(purchasing_intelligence_router)

# SMARTBUY_TRANSFER_INTELLIGENCE_ROUTER
from app.transfer_intelligence import router as transfer_intelligence_router
app.include_router(transfer_intelligence_router)


# SMARTBUY_SPRINT_3_ROUTER
from app.master_data import router as master_data_router

app.include_router(master_data_router)

# SMARTBUY_SPRINT_4_ROUTER
from app.intelligent_products import router as intelligent_products_router

app.include_router(intelligent_products_router)

# SMARTBUY_SPRINT_6_ROUTER
from app.integration_core import router as integration_core_router
from app.integration.eip_api import router as eip_api_router

app.include_router(integration_core_router)
app.include_router(eip_api_router)


# SMARTBUY_EXECUTIVE_COCKPIT_ROUTER
from app.executive_cockpit import router as executive_cockpit_router
app.include_router(executive_cockpit_router)
